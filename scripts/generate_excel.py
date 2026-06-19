"""
Génère data/personnel_medical.xlsx — classeur multi-feuilles simulant
le fichier RH exporté par le système de gestion du personnel hospitalier.

Feuilles :
  - Medecins   : 60 médecins avec spécialité, service, grade
  - Infirmiers : 120 infirmiers avec service, horaire
  - Plannings  : 300 lignes de gardes/astreintes sur 3 mois
"""
import sys, types, random
from datetime import date, timedelta
from pathlib import Path

# Stub lxml pour contourner la version système cassée (Python 3.13)
if "lxml.etree" not in sys.modules:
    _m = types.ModuleType("lxml.etree")
    _m.LXML_VERSION = (0, 0, 0, 0)
    sys.modules.setdefault("lxml", types.ModuleType("lxml"))
    sys.modules["lxml.etree"] = _m

from faker import Faker
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

fake = Faker("fr_FR")
random.seed(42)
Faker.seed(42)

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"

SERVICES = [
    ("S01","Cardiologie"),("S02","Neurologie"),("S03","Urgences"),
    ("S04","Chirurgie générale"),("S05","Pédiatrie"),("S06","Gynécologie"),
    ("S07","Orthopédie"),("S08","Oncologie"),("S09","Réanimation"),
    ("S10","Pneumologie"),("S11","Gastro-entérologie"),("S12","Rhumatologie"),
    ("S13","Psychiatrie"),("S14","Dermatologie"),("S15","Endocrinologie"),
]
HOPITAUX = [f"H{i:02d}" for i in range(1, 21)]
SPECIALITES = [s[1] for s in SERVICES]
GRADES_MED = ["Interne","Chef de clinique","Praticien Hospitalier","PU-PH","MCU-PH"]
GRADES_INF = ["IDE","IADE","IBODE","Cadre de santé","Aide-soignant"]
HORAIRES   = ["Matin (7h-15h)","Après-midi (15h-23h)","Nuit (23h-7h)","Journée (9h-17h)"]
TYPE_GARDE = ["Garde","Astreinte","Repos compensateur","Congé","Formation"]


def _style_header(ws, row=1, fill_color="1F4E79"):
    fill = PatternFill("solid", fgColor=fill_color)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def gen_excel():
    wb = openpyxl.Workbook()

    # ── Feuille 1 : Médecins ──────────────────────────────────────────────────
    ws_med = wb.active
    ws_med.title = "Medecins"
    headers_med = ["medecin_id","nom","prenom","specialite","service_id",
                   "hopital_id","grade","date_recrutement","email","telephone"]
    ws_med.append(headers_med)
    _style_header(ws_med)

    for i in range(1, 61):
        svc   = random.choice(SERVICES)
        recru = date(random.randint(1995, 2023), random.randint(1,12), random.randint(1,28))
        ws_med.append([
            f"MED{i:03d}",
            fake.last_name(),
            fake.first_name(),
            svc[1],
            svc[0],
            random.choice(HOPITAUX),
            random.choice(GRADES_MED),
            recru.isoformat(),
            fake.email(),
            fake.phone_number(),
        ])
    _autofit(ws_med)

    # ── Feuille 2 : Infirmiers ────────────────────────────────────────────────
    ws_inf = wb.create_sheet("Infirmiers")
    headers_inf = ["infirmier_id","nom","prenom","service_id","hopital_id",
                   "grade","horaire","date_recrutement","email"]
    ws_inf.append(headers_inf)
    _style_header(ws_inf, fill_color="1A5276")

    for i in range(1, 121):
        svc   = random.choice(SERVICES)
        recru = date(random.randint(2000, 2023), random.randint(1,12), random.randint(1,28))
        ws_inf.append([
            f"INF{i:03d}",
            fake.last_name(),
            fake.first_name(),
            svc[0],
            random.choice(HOPITAUX),
            random.choice(GRADES_INF),
            random.choice(HORAIRES),
            recru.isoformat(),
            fake.email(),
        ])
    _autofit(ws_inf)

    # ── Feuille 3 : Plannings ─────────────────────────────────────────────────
    ws_plan = wb.create_sheet("Plannings")
    headers_plan = ["planning_id","personnel_id","type_personnel","hopital_id",
                    "service_id","date","type_garde","heure_debut","heure_fin","statut"]
    ws_plan.append(headers_plan)
    _style_header(ws_plan, fill_color="154360")

    start = date(2024, 1, 1)
    for i in range(1, 301):
        d         = start + timedelta(days=random.randint(0, 89))
        is_med    = random.random() < 0.4
        pid       = f"MED{random.randint(1,60):03d}" if is_med else f"INF{random.randint(1,120):03d}"
        type_g    = random.choice(TYPE_GARDE)
        h_debut   = random.choice(["07:00","09:00","15:00","23:00"])
        duree     = 8 if "Nuit" not in type_g else 12
        h_fin_h   = (int(h_debut[:2]) + duree) % 24
        ws_plan.append([
            f"PL{i:04d}",
            pid,
            "Médecin" if is_med else "Infirmier",
            random.choice(HOPITAUX),
            random.choice(SERVICES)[0],
            d.isoformat(),
            type_g,
            h_debut,
            f"{h_fin_h:02d}:00",
            random.choice(["Confirmé","En attente","Annulé"]),
        ])
    _autofit(ws_plan)

    path = DATA_DIR / "personnel_medical.xlsx"
    wb.save(path)
    print(f"  ✓ personnel_medical.xlsx — Médecins: 60, Infirmiers: 120, Plannings: 300")
    return path


if __name__ == "__main__":
    gen_excel()
